import pandas as pd
from datetime import datetime

from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment


def create_report(results):

    report = pd.DataFrame(results)

    total = len(report)

    ok_count = len(
        report[report["Статус"] == "OK"]
    )

    error_count = len(
        report[report["Статус"] == "Ошибка"]
    )

    duplicate_count = len(
        report[report["Тип"] == "Дубликат"]
    )

    missing_1c_count = len(
        report[report["Тип"] == "Нет в 1С"]
    )

    missing_esf_count = len(
        report[report["Тип"] == "Нет в ЭСФ"]
    )

    discrepancy_count = len(
        report[report["Тип"] == "Расхождение"]
    )

    match_percent = round(
        (ok_count / total) * 100,
        2
    ) if total > 0 else 0

    now = datetime.now()

    summary = pd.DataFrame({
        "Показатель": [
            "Дата формирования",
            "Время формирования",
            "Всего записей",
            "Совпало",
            "Ошибок",
            "Расхождений",
            "Дубликатов",
            "Нет в 1С",
            "Нет в ЭСФ",
            "Процент совпадения"
        ],
        "Значение": [
            now.strftime("%d.%m.%Y"),
            now.strftime("%H:%M:%S"),
            total,
            ok_count,
            error_count,
            discrepancy_count,
            duplicate_count,
            missing_1c_count,
            missing_esf_count,
            f"{match_percent}%"
        ]
    })

    filename = "reports/report.xlsx"

    with pd.ExcelWriter(
        filename,
        engine="openpyxl"
    ) as writer:

        summary.to_excel(
            writer,
            sheet_name="Сводка",
            index=False
        )

        report.to_excel(
            writer,
            sheet_name="Проверка",
            index=False
        )

        report[
            report["Тип"] == "Совпадение"
        ].to_excel(
            writer,
            sheet_name="Совпадения",
            index=False
        )

        report[
            report["Тип"] == "Расхождение"
        ].to_excel(
            writer,
            sheet_name="Расхождения",
            index=False
        )

        report[
            report["Тип"] == "Дубликат"
        ].to_excel(
            writer,
            sheet_name="Дубликаты",
            index=False
        )

        report[
            report["Тип"] == "Нет в 1С"
        ].to_excel(
            writer,
            sheet_name="Нет в 1С",
            index=False
        )

        report[
            report["Тип"] == "Нет в ЭСФ"
        ].to_excel(
            writer,
            sheet_name="Нет в ЭСФ",
            index=False
        )

    wb = load_workbook(filename)

    green_fill = PatternFill(
        start_color="C6EFCE",
        end_color="C6EFCE",
        fill_type="solid"
    )

    red_fill = PatternFill(
        start_color="FFC7CE",
        end_color="FFC7CE",
        fill_type="solid"
    )

    blue_fill = PatternFill(
        start_color="D9EAF7",
        end_color="D9EAF7",
        fill_type="solid"
    )

    header_fill = PatternFill(
        start_color="B7DEE8",
        end_color="B7DEE8",
        fill_type="solid"
    )

    bold_font = Font(
        bold=True
    )

    for sheet in wb.worksheets:

        for cell in sheet[1]:
            cell.font = bold_font
            cell.fill = header_fill

    ws_summary = wb["Сводка"]

    for row in ws_summary.iter_rows(
        min_row=2
    ):

        indicator = row[0].value

        if indicator == "Совпало":

            for cell in row:
                cell.fill = green_fill

        elif indicator in [
            "Ошибок",
            "Расхождений",
            "Дубликатов",
            "Нет в 1С",
            "Нет в ЭСФ"
        ]:

            for cell in row:
                cell.fill = red_fill

        else:

            for cell in row:
                cell.fill = blue_fill

    ws_check = wb["Проверка"]

    for row in ws_check.iter_rows(
        min_row=2
    ):

        status = row[1].value

        if status == "OK":

            for cell in row:
                cell.fill = green_fill

        else:

            for cell in row:
                cell.fill = red_fill

    for sheet in wb.worksheets:

        for row in sheet.iter_rows():

            for cell in row:

                cell.alignment = Alignment(
                    vertical="center"
                )

    for sheet in wb.worksheets:

        for column in sheet.columns:

            max_length = 0

            letter = column[0].column_letter

            for cell in column:

                try:

                    max_length = max(
                        max_length,
                        len(str(cell.value))
                    )

                except:
                    pass

            sheet.column_dimensions[
                letter
            ].width = max_length + 5

    for sheet in wb.worksheets:

        sheet.auto_filter.ref = sheet.dimensions

    wb.save(filename)

    print(
        f"Отчёт успешно создан: {filename}"
    )